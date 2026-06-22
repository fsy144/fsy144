import sqlite3
import uuid
import os
import qrcode
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# 配置（修改为官网首页）
DOMAIN = "https://pharmanewzealand.com.cn"  # 改为.cn的官网
QR_SAVE_PATH = "static/qr_codes"
LOGO_PATH = "static/images/logo.jpg"
EXCEL_SAVE_PATH = "."  # Excel文件保存到当前文件夹

# 图片尺寸配置
CANVAS_WIDTH = 750
CANVAS_HEIGHT = 1000
QR_SIZE = 600
LOGO_WIDTH = 600  # 序列号宽度会自动匹配此宽度
TOP_MARGIN = 15
SERIAL_SPACING = 30  # LOGO底部到序列号的间距
SERIAL_TO_QR_SPACING = 60  # 序列号到二维码的间距

if not os.path.exists(QR_SAVE_PATH):
    os.makedirs(QR_SAVE_PATH)


def get_next_serial_no():
    conn = sqlite3.connect('database/anti_fake.db')
    cursor = conn.cursor()
    cursor.execute("SELECT serial_no FROM anti_fake_records ORDER BY serial_no DESC LIMIT 1")
    result = cursor.fetchone()
    conn.close()

    if result:
        last_no = int(result[0])
        next_no = last_no + 1
    else:
        next_no = 1
    return f"{next_no:010d}"


def create_branded_qr_code(url, serial_no, filename):
    canvas = Image.new('RGB', (CANVAS_WIDTH, CANVAS_HEIGHT), color='black')
    draw = ImageDraw.Draw(canvas)

    # 加载并放置LOGO
    logo = Image.open(LOGO_PATH).convert('RGB')
    logo_ratio = logo.height / logo.width
    logo_height = int(LOGO_WIDTH * logo_ratio)
    logo = logo.resize((LOGO_WIDTH, logo_height), Image.Resampling.LANCZOS)
    logo_x = (CANVAS_WIDTH - LOGO_WIDTH) // 2
    logo_y = TOP_MARGIN
    canvas.paste(logo, (logo_x, logo_y))

    # -------------------------- 核心修改：小字体+宽度严格对齐LOGO --------------------------
    serial_text = serial_no
    target_width = LOGO_WIDTH  # 目标宽度与LOGO一致

    # 动态计算合适的字体大小（从较小的初始值开始尝试）
    font_size = 80  # 降低初始字体大小，避免过大
    font_serial = None
    text_width = 0
    text_height = 0

    # 加载字体（保持兼容性）
    while font_size > 20:
        try:
            font_serial = ImageFont.truetype("arial.ttf", font_size)
        except IOError:
            try:
                font_serial = ImageFont.truetype("/Library/Fonts/Arial.ttf", font_size)
            except IOError:
                font_serial = ImageFont.load_default(size=font_size)

        # 获取文本实际尺寸
        bbox = draw.textbbox((0, 0), serial_text, font=font_serial)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        if text_width <= target_width:
            break
        font_size -= 2  # 每次减小2号

    # 计算序列号位置（水平居中，宽度视觉对齐LOGO）
    # 即使文本略小于LOGO，也通过居中保持视觉对齐
    serial_x = (CANVAS_WIDTH - text_width) // 2
    serial_y = logo_y + logo_height + SERIAL_SPACING

    # 绘制白色序列号文本（无红框）
    draw.text((serial_x, serial_y), serial_text, fill="white", font=font_serial)

    # -------------------------- 核心修改结束 --------------------------

    # 生成并放置二维码
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_img = qr_img.resize((QR_SIZE, QR_SIZE))

    qr_x = (CANVAS_WIDTH - QR_SIZE) // 2
    qr_y = serial_y + text_height + SERIAL_TO_QR_SPACING
    canvas.paste(qr_img, (qr_x, qr_y))

    canvas.save(os.path.join(QR_SAVE_PATH, filename))


def insert_into_db(qr_id, serial_no):
    conn = sqlite3.connect('database/anti_fake.db')
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO anti_fake_records (qr_id, serial_no) VALUES (?, ?)", (qr_id, serial_no))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def create_excel_workbook():
    """创建并初始化Excel工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "防伪码生成记录"

    # 设置表头
    headers = ["序号", "防伪编号", "QR ID", "验证链接", "生成时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 20

    return wb, ws


def add_record_to_excel(ws, row_num, serial_no, qr_id, url, generate_time):
    """向Excel添加一条记录"""
    data = [row_num - 1, serial_no, qr_id, url, generate_time]
    for col_num, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def main(num_to_generate=2500):
    print(f"开始生成 {num_to_generate} 个防伪码...")

    # 创建Excel工作簿
    wb, ws = create_excel_workbook()
    current_row = 2  # 从第2行开始写入数据（第1行是表头）

    # 获取当前时间作为Excel文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"防伪码生成记录_{timestamp}.xlsx"
    excel_full_path = os.path.join(EXCEL_SAVE_PATH, excel_filename)

    success_count = 0
    fail_count = 0

    for i in range(num_to_generate):
        qr_id = str(uuid.uuid4())
        serial_no = get_next_serial_no()
        generate_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if insert_into_db(qr_id, serial_no):
            url = f"{DOMAIN}#qr_id={qr_id}"
            filename = f"QR_{serial_no}.png"
            create_branded_qr_code(url, serial_no, filename)

            # 添加到Excel
            add_record_to_excel(ws, current_row, serial_no, qr_id, url, generate_time)
            current_row += 1
            success_count += 1

            print(f"[成功] 编号: {serial_no} | 链接: {url}")
        else:
            fail_count += 1
            print(f"[失败] 编号重复，跳过: {serial_no}")

    # 保存Excel文件
    wb.save(excel_full_path)

    # 打印统计信息
    print("\n" + "=" * 50)
    print(f"生成完成！")
    print(f"成功生成: {success_count} 个")
    print(f"失败跳过: {fail_count} 个")
    print(f"Excel记录已保存至: {os.path.abspath(excel_full_path)}")
    print("=" * 50)


if __name__ == "__main__":
    # 首先安装依赖
    # pip install openpyxl pillow qrcode
    main(40000)